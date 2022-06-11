from datetime import datetime
import json
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import FileResponse, JsonResponse, HttpResponse
from django.db.models.functions import TruncMonth
from .models import Transaction, Loan, Collection
from django.db.models import Sum
from .utils import generate_to_pdf
from num2words import num2words
from dateutil.relativedelta import relativedelta
import xlsxwriter
from django.conf import settings
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill

@ensure_csrf_cookie
def set_csrf_token(request):
    return JsonResponse({"details": "CSRF cookie set"})

@require_POST
def login_view(request):
    data = json.loads(request.body)
    username = data.get('username')
    password = data.get('password')
    user = authenticate(username=username, password=password)

    if user is not None:
        login(request, user)
        return JsonResponse({"detail": "Success"})

    return JsonResponse({"detail": "Invalid credentials"}, status=400)

@require_POST
def logout_view(request):
    logout(request)
    return JsonResponse({"detail": "Success"})

def check_session(request):
    print(request.user.is_authenticated)
    if request.user.is_authenticated:
        return JsonResponse({"user": request.user.username})

    return JsonResponse({"detail": "No session user"}, status=400)

def cash_flow_statement(request, start_date, end_date=None):
    report_title = "Daily Cash Flow Statement"
    ranged = False

    transactions = Transaction.objects.filter(
        post_date=start_date
    ).order_by('post_date')

    if end_date is not None:
        ranged = True
        report_title = "Ranged Cash Flow Statement"
        transactions = Transaction.objects.filter(
            post_date__gte=start_date,
            post_date__lte=end_date
        ).order_by('post_date')

    total_debit = transactions.filter(transaction_side="debit").aggregate(
        Sum('amount')
    )
    total_credit = transactions.filter(transaction_side="credit").aggregate(
        Sum('amount')
    )

    context = {
        'transactions': transactions,
        'report_title': report_title,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'ranged': ranged,
        'start_date': start_date,
        'end_date': end_date
    }

    pdf = generate_to_pdf("cash_flow_statement.html", context, f"cash-flow-statement-{start_date.strftime('%Y-%m-%d')}")

    return FileResponse(open(pdf, 'rb'), content_type="application/pdf")

def computation_report(request, loan_id):
    loan = Loan.objects.get(pk=loan_id)

    total_deductions = round(loan.udi + loan.llrf + loan.processing_fee + loan.fee_others, 2)
    installment = loan.principal_amount / loan.term
    principal_amount = loan.principal_amount

    if not loan.is_advance:
        total_deductions = round(loan.llrf + loan.processing_fee + loan.fee_others, 2)
        principal_amount = loan.principal_amount + loan.udi
        installment = principal_amount / loan.term

    context = {
        'loan': loan,
        'total_loan_amount': loan.principal_amount + loan.udi,
        'total_deductions': total_deductions,
        'installment': installment
    }

    pdf = generate_to_pdf("computation_report.html", context, f"computation-report-{loan_id}")

    return FileResponse(open(pdf, 'rb'), content_type="application/pdf")

def promissory_report(request, loan_id):
    loan = Loan.objects.get(pk=loan_id)

    total_deductions = round(loan.udi + loan.llrf + loan.processing_fee + loan.fee_others, 2)
    installment = loan.principal_amount / loan.term
    principal_amount = loan.principal_amount

    if not loan.is_advance:
        total_deductions = round(loan.llrf + loan.processing_fee + loan.fee_others, 2)
        principal_amount = loan.principal_amount + loan.udi
        installment = principal_amount / loan.term

    context = {
        'loan': loan,
        'total_loan_amount': loan.principal_amount + loan.udi,
        'total_deductions': total_deductions,
        'installment': installment,
        'principal_amount_words': num2words(loan.principal_amount).upper(),
        'loan_detail': loan.loan_detail_set.all(),
        'name_one': request.GET.get('name_one', ''),
        'name_two': request.GET.get('name_two', ''),
        'name_three': request.GET.get('name_three', '')
    }

    pdf = generate_to_pdf("promissory_report.html", context, f"promissory-report-{loan_id}", page_size='Legal')

    return FileResponse(open(pdf, 'rb'), content_type="application/pdf")

def disclosure_of_loan(request, loan_id):
    loan = Loan.objects.get(pk=loan_id)
    
    non_financial_charge_total = loan.llrf + loan.processing_fee + loan.fee_others

    installment = loan.principal_amount / loan.term
    principal_amount = loan.principal_amount

    if not loan.is_advance:
        principal_amount = loan.principal_amount + loan.udi
        installment = principal_amount / loan.term

    context = {
        'loan': loan,
        'non_financial_charge_total': non_financial_charge_total,
        'installment': installment
    }

    # or Folio?
    pdf = generate_to_pdf("disclosure_of_loan.html", context, f"disclosure-form-{loan_id}", page_size='Legal')
    return FileResponse(open(pdf, 'rb'), content_type="application/pdf")

def retrieve_opening_cash_balance(request, start_date):
    # Fetch Transactions Before start_date
    opening_balance = 0

    debit_transactions = Transaction.objects.filter(
        post_date__lt=start_date,
        transaction_side='debit'
    ).aggregate(Sum('amount'))
    credit_transactions = Transaction.objects.filter(
        post_date__lt=start_date,
        transaction_side='credit'
    ).aggregate(Sum('amount'))

    opening_balance = (debit_transactions['amount__sum'] or 0) - (credit_transactions['amount__sum'] or 0)

    side = 'debit'

    if opening_balance < 0:
        side = 'credit'

    return JsonResponse({'opening_balance': abs(opening_balance), 'side': side}, safe=False)

def loan_masterlist(request, year, month):
    loan_date = datetime.now().replace(year=year, month=month, day=1) + relativedelta(months=1)
    loans = Loan.objects.filter(date_granted__lt=loan_date)

    # We just want the monthly transactions
    offset_date = loan_date.replace(month=month)
    transactions = loans.filter(date_granted__gte=offset_date).order_by('date_granted')

    aggregates = loans.annotate(
        month=TruncMonth('date_granted')).values('month').order_by('-month').annotate(
            total_principal_loan=Sum('principal_amount'),
            total_udi=Sum('udi'),
            total_cashout=Sum('net_cash_out'),
            total_llrf=Sum('llrf'),
            total_pfee=Sum('processing_fee'),
            total_others=Sum('fee_others')
        )

    totals = loans.aggregate(
        total_principal_loan=Sum('principal_amount'),
        total_udi=Sum('udi'),
        total_cashout=Sum('net_cash_out'),
        total_llrf=Sum('llrf'),
        total_pfee=Sum('processing_fee'),
        total_others=Sum('fee_others')
    )

    context = {
        'aggregates': aggregates,
        'totals': totals,
        'loan_date': loan_date.replace(month=month),
        'transactions': transactions
    }

    # or Folio?
    pdf = generate_to_pdf("loan_masterlist.html", context, f"loan-masterlist-{year}", page_size='Legal', orientation="landscape")
    return FileResponse(open(pdf, 'rb'), content_type="application/pdf")

def cashloan_masterlist(request, year, month):
    loan_date = datetime.now().replace(year=year, month=month, day=1) + relativedelta(months=1)
    loans = Loan.objects.filter(date_granted__lt=loan_date, is_cash_advance=True)

    # We just want the monthly transactions
    offset_date = loan_date.replace(month=month)
    transactions = loans.filter(date_granted__gte=offset_date).order_by('date_granted')

    aggregates = loans.annotate(
        month=TruncMonth('date_granted')).values('month').order_by('-month').annotate(
            total_principal_loan=Sum('principal_amount'),
            total_udi=Sum('udi'),
            total_cashout=Sum('net_cash_out'),
            total_llrf=Sum('llrf'),
            total_pfee=Sum('processing_fee'),
            total_others=Sum('fee_others')
        )

    totals = loans.aggregate(
        total_principal_loan=Sum('principal_amount'),
        total_udi=Sum('udi'),
        total_cashout=Sum('net_cash_out'),
        total_llrf=Sum('llrf'),
        total_pfee=Sum('processing_fee'),
        total_others=Sum('fee_others')
    )

    context = {
        'aggregates': aggregates,
        'totals': totals,
        'loan_date': loan_date.replace(month=month),
        'transactions': transactions
    }

    # or Folio?
    pdf = generate_to_pdf("loan_masterlist.html", context, f"cashloan-masterlist-{year}", page_size='Legal', orientation="landscape")
    return FileResponse(open(pdf, 'rb'), content_type="application/pdf")

def cash_receipts_xls(request, year, month):
    path = f"{settings.REPORTS_PATH}/cash-receipts-{month}-{year}.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = f'{month}'

    collections = Collection.objects.filter(post_date__month=month, post_date__year=year).order_by('post_date__month', 'post_date__day').values(
        'total_amount_to_pay',
        'client__first_name',
        'client__last_name',
        'client__bank_name',
        'collection_amount',
        'refundable_amount',
        'post_date')

    # styles
    white_fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type = "solid")
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False)

    last_date = None
    ws._current_row = 1
    table_open_row = None
    table_close_row = None


    for index, collection in enumerate(collections):
        current_date = collection['post_date']

        if last_date != current_date:
            table_close_row = ws._current_row

            # Close and add table
            if table_open_row != None:
                table = Table(displayName=f"CashReceipt-{table_open_row}", ref=f"A{table_open_row}:I{table_close_row}")
                table.tableStyleInfo = table_style
                ws.add_table(table)

            # Recreate Headers
            if ws._current_row != 1:
                ws._current_row += 5

            header = ws.cell(row=ws._current_row, column=1, value=f"D'Last Frontier Lending Corporation\nCASH RECEIPTS\n{current_date.strftime('%B %d, %Y')}")
            header.fill = white_fill
            ws.merge_cells(f'A{ws._current_row}:I{ws._current_row}')
            ws._current_row += 1
            ws.append(["NAME", "BANKS", "AMOUNT", "LOAN", "AR\n(CLIENT)", "AR\n(PERSONNEL)", "AR\n(OTHERS)", "AP", "INTEREST"])
            table_open_row = ws._current_row

        ar_to_client = 0
        if collection['collection_amount'] < collection['total_amount_to_pay']:
            ar_to_client = collection['total_amount_to_pay'] - collection['collection_amount']

        ws.append([
            f"{collection['client__last_name']}, {collection['client__first_name']}",
            collection['client__bank_name'],
            f"{collection['collection_amount']:,.2f}",
            f"{collection['total_amount_to_pay']:,.2f}",
            f"{ar_to_client:,.2f}",
            "",
            "",
            f"{collection['refundable_amount'] or 0:,.2f}",
            ""
        ])

        if index == len(collections) - 1:
            table = Table(displayName=f"CashReceipt-{table_open_row}", ref=f"A{table_open_row}:I{ws._current_row}")
            table.tableStyleInfo = table_style
            ws.add_table(table)

        last_date = current_date

    wb.save(path)
    return FileResponse(open(path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def loan_masterlist_xls(request, year, month):
    path = f"{settings.REPORTS_PATH}/loan-masterlist-{month}-{year}.xlsx"
    loan_date = datetime.now().replace(year=year, month=month, day=1) + relativedelta(months=1)
    loans = Loan.objects.filter(date_granted__lt=loan_date, loan_type='pension')

    # We just want the monthly transactions
    offset_date = loan_date.replace(month=month)
    transactions = loans.filter(date_granted__gte=offset_date).order_by('date_granted')

    aggregates = loans.annotate(
        month=TruncMonth('date_granted')).values('month').order_by('-month').annotate(
            total_principal_loan=Sum('principal_amount'),
            total_udi=Sum('udi'),
            total_cashout=Sum('net_cash_out'),
            total_llrf=Sum('llrf'),
            total_pfee=Sum('processing_fee'),
            total_others=Sum('fee_others')
        )

    totals = loans.aggregate(
        total_principal_loan=Sum('principal_amount'),
        total_udi=Sum('udi'),
        total_cashout=Sum('net_cash_out'),
        total_llrf=Sum('llrf'),
        total_pfee=Sum('processing_fee'),
        total_others=Sum('fee_others')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}'

    # styles
    green_fill = PatternFill(start_color="0099CC00", end_color="0099CC00", fill_type = "solid")
    white_fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type = "solid")

    header1 = ws.cell(row=1, column=1, value=f"D'Last Frontier Lending Corporation\nSSS PENSIONER'S LOAN MASTERLIST\n{offset_date.strftime('%B, %Y')}")
    header1.fill = white_fill
    ws.merge_cells('A1:L1')

    ws.append(["Date", "Name of Pensioner", "CTRL No", "Term", "Principal Loan", "UDI", "Cash-out", "LLRF", "P/FEE", "VAT", "Misc. Income", "UDI Rebates"])

    last_date = None
    for row in transactions:
        current_date = row.date_granted

        if last_date != current_date:
            show_date = row.date_granted.strftime("%b %d")
        else:
            show_date = ""

        last_date = current_date

        ws.append([
            show_date,
            f"{row.client.last_name}, {row.client.first_name}",
            row.control_number,
            row.term,
            f"{row.principal_amount:,.2f}",
            f"{row.udi:,.2f}",
            f"{row.net_cash_out:,.2f}",
            f"{row.llrf:,.2f}",
            f"{row.processing_fee:,.2f}",
            f"{row.fee_others:,.2f}",
            "",
            ""
        ])

    table = Table(displayName="LoanMasterlistTable", ref=f"A2:L{transactions.count() + 2}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    
    table.tableStyleInfo = style

    ws.add_table(table)

    row_counter = transactions.count() + 3

    for aggregate in aggregates:
        _ = ws.cell(row=row_counter, column=1, value=f"{aggregate['month'].strftime('%B, %Y')}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=2, value="")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=3, value="")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=4, value="")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=5, value=f"{aggregate['total_principal_loan']:,.2f}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=6, value=f"{aggregate['total_udi']:,.2f}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=7, value=f"{aggregate['total_cashout']:,.2f}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=8, value=f"{aggregate['total_llrf']:,.2f}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=9, value=f"{aggregate['total_pfee']:,.2f}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=10, value=f"{aggregate['total_others']:,.2f}")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=11, value="")
        _.fill = green_fill
        _ = ws.cell(row=row_counter, column=12, value="")
        _.fill = green_fill

        row_counter += 1

    # totals
    _ = ws.cell(row=row_counter, column=1, value="TOTALS")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=2, value="")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=3, value="")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=4, value="")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=5, value=f"{totals['total_principal_loan']:,.2f}")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=6, value=f"{totals['total_udi']:,.2f}")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=7, value=f"{totals['total_cashout']:,.2f}")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=8, value=f"{totals['total_llrf']:,.2f}")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=9, value=f"{totals['total_pfee']:,.2f}")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=10, value=f"{totals['total_others']:,.2f}")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=11, value="")
    _.fill = green_fill
    _ = ws.cell(row=row_counter, column=12, value="")
    _.fill = green_fill

    wb.save(f"{path}")

    return FileResponse(open(path, 'rb'), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
